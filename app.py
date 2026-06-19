"""
Pharma HTML Report → Excel Converter
S.F. Medical Agency — FoxPro/Tally fixed-width absolute-positioned HTML parser

Parsing strategy:
  - HTML has multiple <div class="page"> blocks (one per printed page)
  - Each page uses position:relative, so top values reset per page
  - All content divs are sorted by top position WITHIN each page
  - Item rows use fixed-width \xa0 (non-breaking space) padding
  - Last 5 tokens of each item row = Qty, Free, Rate, Amount, %
  - '-' in Free position = 0 (blank in output)
"""

import io
import re
from flask import Flask, request, jsonify, send_file, render_template
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32 MB

# ─────────────────────────────────────────────
#  TEXT NORMALISATION
# ─────────────────────────────────────────────

NBSP  = '\xa0'       # non-breaking space used as column padding
NBHY  = '\u2011'     # non-breaking hyphen used in item names
RSQUO = '\u2019'     # right single quote (SPECIALITY'S)


def normalize(raw: str) -> str:
    """Convert non-breaking chars to standard ASCII equivalents."""
    return (
        raw.replace(NBSP, ' ')
           .replace(NBHY, '-')
           .replace(RSQUO, "'")
           .strip()
    )


# ─────────────────────────────────────────────
#  LINE CLASSIFICATION
# ─────────────────────────────────────────────

# Phrases that mark non-data lines (header, footer, page info)
SKIP_PHRASES = [
    'S.F.MEDICAL', 'PARTY / ITEM', 'PARTY/ ITEM', 'Report For',
    'Company', 'D E S C R I P T I O N', 'Continued..', 'Page No',
    'GSTIN', 'Phone', 'End of Report', 'GRAND TOTAL',
]

_NUM_PAT  = re.compile(r'^-?\d+\.?\d*$')
_DASH_PAT = re.compile(r'^-$')
_SEP_PAT  = re.compile(r'^[-\u2011=\s]+$')


def is_separator(text: str) -> bool:
    return bool(_SEP_PAT.match(text))


def is_skip_line(text: str) -> bool:
    return any(p in text for p in SKIP_PHRASES)


def is_party_line(text: str) -> bool:
    """
    Party names are ALL-CAPS lines with NO numeric tokens.
    Rules:
      - Not a separator (-------)
      - No standalone numeric tokens
      - ≥85% uppercase alpha characters
    """
    if is_separator(text):
        return False
    tokens = text.split()
    if not tokens:
        return False
    for tok in tokens:
        if _NUM_PAT.match(tok):          # numeric token → item row, not party
            return False
    alpha = [c for c in text if c.isalpha()]
    if not alpha:
        return False
    upper_ratio = sum(1 for c in alpha if c.isupper()) / len(alpha)
    return upper_ratio >= 0.85


# ─────────────────────────────────────────────
#  ITEM ROW PARSER
# ─────────────────────────────────────────────

def parse_item_row(raw_text: str) -> dict | None:
    """
    Parse one item row from its raw HTML text (with \xa0 padding).

    The row ends with 5 fixed tokens: Qty  Free  Rate  Amount  Pct
    Everything before those 5 tokens = Item Name (name + pack size joined).
    '-' in the Free slot means zero (stored as blank).
    """
    text   = normalize(raw_text)
    tokens = text.split()

    if not tokens:
        return None

    # Collect the last 5 numeric-or-dash tokens from the right
    trailing = []
    i = len(tokens) - 1
    while i >= 0 and len(trailing) < 5:
        tok = tokens[i]
        if _NUM_PAT.match(tok) or _DASH_PAT.match(tok):
            trailing.insert(0, tok)
            i -= 1
        else:
            break

    if len(trailing) < 5:
        return None

    qty_tok, free_tok, rate_tok, amt_tok = (
        trailing[0], trailing[1], trailing[2], trailing[3]
    )
    # trailing[4] is the % column — we don't need it

    try:
        qty    = float(qty_tok)
        free   = None if free_tok == '-' else float(free_tok)
        rate   = float(rate_tok)
        amount = float(amt_tok)
    except ValueError:
        return None

    # Item name = all tokens before the 5 trailing ones
    name_tokens = tokens[: len(tokens) - 5]
    item_name   = ' '.join(name_tokens).strip()

    if not item_name:
        return None

    return {
        'item_name': item_name,
        'qty':       qty,
        'free':      free,      # None means blank (was '-')
        'rate':      rate,
        'amount':    amount,
    }


# ─────────────────────────────────────────────
#  MAIN HTML PARSER
# ─────────────────────────────────────────────

def get_top(div) -> float:
    m = re.search(r'top\s*:\s*([\d.]+)', div.get('style', ''))
    return float(m.group(1)) if m else 0.0


def parse_html(html_content: str) -> list[dict]:
    """
    Parse the FoxPro/Tally HTML report.

    The HTML has 4+ <div class="page"> blocks. Because each page uses
    position:relative, top values reset to 0 at each new page. We must
    process each page independently (sort by top within the page) and
    concatenate results.
    """
    soup  = BeautifulSoup(html_content, 'html.parser')
    pages = soup.find_all('div', class_='page')

    rows          = []
    current_party = None

    for page in pages:
        # Get all content divs inside this page, sorted by top position
        divs = sorted(page.find_all('div', style=True), key=get_top)

        for div in divs:
            raw  = div.get_text()           # keep raw \xa0 for parser
            text = normalize(raw)

            if not text:
                continue
            if is_separator(text):
                continue
            if is_skip_line(text):
                continue
            if 'TOTAL' in text:             # TOTAL / GRAND TOTAL lines → skip
                continue

            if is_party_line(text):
                current_party = text
                continue

            if current_party:
                parsed = parse_item_row(raw)
                if parsed:
                    rows.append({
                        'Party Name': current_party,
                        'Item Name':  parsed['item_name'],
                        'Qty':        parsed['qty'],
                        'Free':       parsed['free'],   # None = blank
                        'Rate':       parsed['rate'],
                        'Amount':     parsed['amount'],
                    })

    return rows


# ─────────────────────────────────────────────
#  EXCEL BUILDER  — matches expected output format
# ─────────────────────────────────────────────

def build_excel(rows: list[dict]) -> bytes:
    """
    Build a formatted .xlsx that matches the expected output format exactly:
    Party Name | Item Name | Qty | Free | Rate | Amount
    Free is blank (not 0) when the source had '-'.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # ── Styles ──────────────────────────────────
    hdr_font   = Font(name='Arial', bold=True,  color='FFFFFF', size=11)
    hdr_fill   = PatternFill('solid', start_color='1B4F72')
    party_font = Font(name='Arial', bold=True,  color='1B2631', size=10)
    party_fill = PatternFill('solid', start_color='D6EAF8')
    data_font  = Font(name='Arial', size=10)
    alt_fill   = PatternFill('solid', start_color='F8FBFD')
    thin       = Side(style='thin', color='CCCCCC')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    c_align    = Alignment(horizontal='center', vertical='center')
    l_align    = Alignment(horizontal='left',   vertical='center')
    r_align    = Alignment(horizontal='right',  vertical='center')

    # ── Header row ──────────────────────────────
    headers    = ['Party Name', 'Item Name', 'Qty', 'Free', 'Rate', 'Amount']
    col_widths = [30, 38, 10, 10, 12, 14]

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell            = ws.cell(row=1, column=ci, value=hdr)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.border     = border
        cell.alignment  = c_align
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    # ── Data rows ───────────────────────────────
    row_num    = 2
    prev_party = None
    alt        = False

    for record in rows:
        party = record['Party Name']

        # Party separator row when party changes
        if party != prev_party:
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num,   end_column=6
            )
            pc            = ws.cell(row=row_num, column=1, value=party)
            pc.font       = party_font
            pc.fill       = party_fill
            pc.border     = border
            pc.alignment  = l_align
            ws.row_dimensions[row_num].height = 18
            row_num   += 1
            prev_party = party
            alt        = False

        fill = PatternFill('solid', start_color='FFFFFF') if not alt else alt_fill

        # Free: blank cell if None, numeric if has value
        free_val = record['Free']   # None → blank, float → number

        values = [
            record['Party Name'],
            record['Item Name'],
            record['Qty'],
            free_val if free_val is not None else '',
            record['Rate'],
            record['Amount'],
        ]
        aligns = [l_align, l_align, c_align, c_align, r_align, r_align]
        fmts   = [None, None, '#,##0', '#,##0', '#,##0.00', '#,##0.00']

        for ci, (val, aln, fmt) in enumerate(zip(values, aligns, fmts), start=1):
            cell           = ws.cell(row=row_num, column=ci, value=val)
            cell.font      = data_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = aln
            if fmt and val != '':
                cell.number_format = fmt

        ws.row_dimensions[row_num].height = 16
        row_num += 1
        alt = not alt

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
#  FLASK ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/parse', methods=['POST'])
def parse_files():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files uploaded'}), 400

    all_rows   = []
    file_stats = []

    for f in files:
        if not f.filename:
            continue
        try:
            html_content = f.read().decode('windows-1252', errors='replace')
            rows = parse_html(html_content)
            all_rows.extend(rows)
            file_stats.append({
                'name':    f.filename,
                'rows':    len(rows),
                'parties': len(set(r['Party Name'] for r in rows)),
                'status':  'ok'
            })
        except Exception as e:
            file_stats.append({
                'name': f.filename, 'rows': 0, 'parties': 0,
                'status': f'error: {str(e)}'
            })

    if not all_rows:
        return jsonify({
            'error': 'No data could be extracted. Check the file format.',
            'file_stats': file_stats
        }), 422

    # Serialise for JSON: convert None → None (JSON null) for Free
    serialisable = []
    for r in all_rows:
        serialisable.append({
            'Party Name': r['Party Name'],
            'Item Name':  r['Item Name'],
            'Qty':        r['Qty'],
            'Free':       r['Free'],      # None or float
            'Rate':       r['Rate'],
            'Amount':     r['Amount'],
        })

    return jsonify({
        'total_rows':    len(serialisable),
        'total_parties': len(set(r['Party Name'] for r in serialisable)),
        'file_stats':    file_stats,
        'preview':       serialisable[:200],
        'data':          serialisable,
    })


@app.route('/download', methods=['POST'])
def download():
    payload  = request.get_json(force=True)
    rows     = payload.get('data', [])
    filename = payload.get('filename', 'pharma_report')

    if not rows:
        return jsonify({'error': 'No data to export'}), 400

    # Restore None for Free (JSON null → None)
    for r in rows:
        if r.get('Free') is None:
            r['Free'] = None

    xlsx_bytes = build_excel(rows)
    buf = io.BytesIO(xlsx_bytes)
    buf.seek(0)
    safe = re.sub(r'[^\w\-.]', '_', filename)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{safe}.xlsx'
    )


if __name__ == '__main__':
    app.run(debug=True, port=5050)
